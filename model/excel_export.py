"""
Excel Export Module
Exports financial models to professional Excel format with formulas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports financial models to Excel with industry-standard formatting
    """
    
    # Color codes following industry standards
    BLUE = '0000FF'  # Inputs
    BLACK = '000000'  # Formulas
    GREEN = '008000'  # Links
    YELLOW = 'FFFF00'  # Key assumptions
    
    def __init__(self):
        self.wb = None
        
    def export_model(self, model):
        """
        Export complete financial model with all scenarios
        """
        self.wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create cover sheet
        self._create_cover_sheet(model)
        
        # Export each scenario
        for scenario in model.scenarios.all():
            self._export_scenario_sheets(scenario)
        
        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_scenario(self, scenario):
        """
        Export single scenario to Excel
        """
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        self._export_scenario_sheets(scenario)
        
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_cover_sheet(self, model):
        """Create cover/summary sheet"""
        ws = self.wb.create_sheet('Cover', 0)
        
        # Title
        ws['A1'] = model.name
        ws['A1'].font = Font(size=20, bold=True)
        
        # Metadata
        ws['A3'] = 'Project Type:'
        ws['B3'] = model.get_project_type_display()
        ws['A4'] = 'Created:'
        ws['B4'] = model.created_at.strftime('%Y-%m-%d')
        ws['A5'] = 'Last Updated:'
        ws['B5'] = model.updated_at.strftime('%Y-%m-%d')
        ws['A6'] = 'Owner:'
        ws['B6'] = model.owner.email
        
        # Scenarios
        ws['A8'] = 'Scenarios:'
        row = 9
        for scenario in model.scenarios.all():
            ws[f'A{row}'] = f"  • {scenario.name}"
            row += 1
        
        # Format
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _export_scenario_sheets(self, scenario):
        """Create all sheets for a scenario"""
        prefix = scenario.name[:10]  # Limit sheet name length
        
        # Create Assumptions Sheet
        self._create_assumptions_sheet(scenario, prefix)
        
        # Create Financial Statements
        self._create_income_statement(scenario, prefix)
        self._create_balance_sheet(scenario, prefix)
        self._create_cashflow_statement(scenario, prefix)
        
        # Create Supporting Schedules
        self._create_revenue_schedule(scenario, prefix)
        self._create_opex_schedule(scenario, prefix)
        self._create_depreciation_schedule(scenario, prefix)
        self._create_debt_schedule(scenario, prefix)
        
        # Create Ratios & Valuation
        self._create_ratios_sheet(scenario, prefix)
        self._create_valuation_sheet(scenario, prefix)
    
    def _create_assumptions_sheet(self, scenario, prefix):
        """Create comprehensive assumptions sheet"""
        ws = self.wb.create_sheet(f'{prefix}_Assumptions')
        
        row = 1
        
        # Header
        ws[f'A{row}'] = 'ASSUMPTIONS'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Project Information
        if hasattr(scenario, 'project_info'):
            pi = scenario.project_info
            ws[f'A{row}'] = 'PROJECT INFORMATION'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            self._add_assumption(ws, row, 'Project Name', pi.project_name)
            row += 1
            self._add_assumption(ws, row, 'Location', pi.project_location)
            row += 1
            self._add_assumption(ws, row, 'Capacity', pi.total_capacity, pi.capacity_unit)
            row += 2
        
        # Macro Assumptions
        if hasattr(scenario, 'macro_assumptions'):
            macro = scenario.macro_assumptions
            ws[f'A{row}'] = 'MACRO ECONOMIC'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            self._add_assumption(ws, row, 'Currency', macro.reporting_currency)
            row += 1
            self._add_assumption(ws, row, 'Exchange Rate', macro.exchange_rate_local_per_usd)
            row += 1
            self._add_assumption(ws, row, 'Local Inflation', macro.local_inflation_rate, '%')
            row += 1
            self._add_assumption(ws, row, 'Discount Rate', macro.discount_rate_wacc, '%')
            row += 2
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def _create_income_statement(self, scenario, prefix):
        """Create Income Statement"""
        ws = self.wb.create_sheet(f'{prefix}_IS')
        
        # Get calculated data
        statements = scenario.calculated_statements.filter(statement_type='is')
        
        if not statements.exists():
            ws['A1'] = 'No calculated data available. Please run calculation first.'
            return
        
        # Build header
        row = 1
        ws[f'A{row}'] = 'INCOME STATEMENT'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Period headers
        ws[f'A{row}'] = 'Line Item'
        periods = []
        col = 2
        first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
        for period in sorted(first_stmt.values_by_period.keys()):
            ws.cell(row, col).value = period
            ws.cell(row, col).font = Font(bold=True)
            periods.append(period)
            col += 1
        row += 1
        
        # Add line items
        for stmt in statements:
            ws.cell(row, 1).value = stmt.line_item
            col = 2
            for period in periods:
                value = stmt.values_by_period.get(period, 0)
                ws.cell(row, col).value = float(value)
                ws.cell(row, col).number_format = '#,##0.00'
                col += 1
            row += 1
        
        # Format
        self._format_statement_sheet(ws, len(periods))
    
    def _create_balance_sheet(self, scenario, prefix):
        """Create Balance Sheet"""
        ws = self.wb.create_sheet(f'{prefix}_BS')
        
        statements = scenario.calculated_statements.filter(statement_type='bs')
        
        if not statements.exists():
            ws['A1'] = 'No calculated data available.'
            return
        
        # Similar structure to IS
        self._populate_statement_sheet(ws, 'BALANCE SHEET', statements)
        first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
        self._format_statement_sheet(ws, len(first_stmt.values_by_period.keys()))
    
    def _create_cashflow_statement(self, scenario, prefix):
        """Create Cash Flow Statement"""
        ws = self.wb.create_sheet(f'{prefix}_CFS')
        
        statements = scenario.calculated_statements.filter(statement_type='cfs')
        
        if not statements.exists():
            ws['A1'] = 'No calculated data available.'
            return
        
        self._populate_statement_sheet(ws, 'CASH FLOW STATEMENT', statements)
        first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
        self._format_statement_sheet(ws, len(first_stmt.values_by_period.keys()))
    
    def _create_revenue_schedule(self, scenario, prefix):
        """Create detailed revenue schedule"""
        ws = self.wb.create_sheet(f'{prefix}_Revenue')
        
        ws['A1'] = 'REVENUE SCHEDULE'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Pull revenue line items from IS statements
        from .models import CalculatedStatement
        stmts = CalculatedStatement.objects.filter(
            scenario=scenario, statement_type='is',
            line_item__in=['Total Revenue']
        )
        # Also add individual product revenue if available
        all_revenue_stmts = list(stmts)
        
        # Add revenue product details
        row = 3
        ws[f'A{row}'] = 'REVENUE PRODUCTS'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for product in scenario.revenue_products.all():
            ws[f'A{row}'] = product.product_name
            ws[f'B{row}'] = f'Volume: {product.year_1_sales_volume:,.0f} {product.unit_of_measure}'
            ws[f'C{row}'] = f'Price: ${product.unit_price_year_1:,.2f}'
            ws[f'D{row}'] = f'Growth: {product.volume_growth_rate}%'
            ws[f'E{row}'] = f'Escalation: {product.price_escalation_rate}%'
            ws[f'B{row}'].font = Font(color=self.BLUE)
            ws[f'C{row}'].font = Font(color=self.BLUE)
            row += 1
        
        row += 1
        
        # Add total revenue time series if calculated
        if all_revenue_stmts:
            self._populate_statement_sheet(ws, 'CALCULATED REVENUE', all_revenue_stmts)
        
        # Format
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
    
    def _create_opex_schedule(self, scenario, prefix):
        """Create OpEx schedule"""
        ws = self.wb.create_sheet(f'{prefix}_OpEx')
        ws['A1'] = 'OPERATING EXPENSES'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add OpEx assumptions
        row = 3
        try:
            opex = scenario.operating_expenses
            assumptions = [
                ('Total Headcount', opex.total_headcount, ''),
                ('Average Annual Salary', opex.average_annual_salary, '$'),
                ('Benefits/Payroll Tax', opex.benefits_payroll_tax_pct, '%'),
                ('Salary Escalation Rate', opex.salary_escalation_rate, '%'),
                ('Power/Electricity (Annual)', opex.power_electricity_cost_annual, '$'),
                ('Insurance (Annual)', opex.insurance_annual, '$'),
                ('Marketing/Sales (% Rev)', opex.marketing_sales_pct_revenue, '%'),
                ('Regular Maintenance (% Rev)', opex.regular_maintenance_pct_revenue, '%'),
            ]
            ws[f'A{row}'] = 'OPEX ASSUMPTIONS'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for label, value, unit in assumptions:
                self._add_assumption(ws, row, label, value, unit)
                row += 1
        except Exception:
            pass
        
        row += 1
        
        # Add calculated OpEx statements (Staff Costs, Utilities, etc.)
        from .models import CalculatedStatement
        opex_stmts = CalculatedStatement.objects.filter(
            scenario=scenario, statement_type='is',
            line_item__in=['Total Operating Expenses']
        )
        if opex_stmts.exists():
            self._populate_statement_sheet(ws, 'CALCULATED OPEX', opex_stmts)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_depreciation_schedule(self, scenario, prefix):
        """Create depreciation schedule"""
        ws = self.wb.create_sheet(f'{prefix}_Depreciation')
        ws['A1'] = 'DEPRECIATION SCHEDULE'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add depreciation assumptions
        row = 3
        ws[f'A{row}'] = 'ASSET CATEGORIES'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Category'
        ws[f'B{row}'] = 'Asset Value'
        ws[f'C{row}'] = 'Useful Life (yrs)'
        ws[f'D{row}'] = 'Residual Value %'
        ws[f'E{row}'] = 'Method'
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col_letter}{row}'].font = Font(bold=True)
        row += 1
        
        for schedule in scenario.depreciation_schedules.all():
            ws[f'A{row}'] = schedule.get_asset_category_display()
            ws[f'B{row}'] = float(schedule.asset_value)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].font = Font(color=self.BLUE)
            ws[f'C{row}'] = schedule.useful_life_years
            ws[f'D{row}'] = float(schedule.residual_value_pct)
            ws[f'E{row}'] = schedule.get_depreciation_method_display()
            row += 1
        
        # Format
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
    
    def _create_debt_schedule(self, scenario, prefix):
        """Create debt schedule"""
        ws = self.wb.create_sheet(f'{prefix}_Debt')
        
        statements = scenario.calculated_statements.filter(statement_type='debt')
        
        if not statements.exists():
            return
        
        self._populate_statement_sheet(ws, 'DEBT SCHEDULE', statements)
        first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
        self._format_statement_sheet(ws, len(first_stmt.values_by_period.keys()))
    
    def _create_ratios_sheet(self, scenario, prefix):
        """Create financial ratios sheet"""
        ws = self.wb.create_sheet(f'{prefix}_Ratios')
        
        statements = scenario.calculated_statements.filter(statement_type='ratio')
        
        if statements.exists():
            self._populate_statement_sheet(ws, 'FINANCIAL RATIOS', statements)
            first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
            self._format_statement_sheet(ws, len(first_stmt.values_by_period.keys()))
    
    def _create_valuation_sheet(self, scenario, prefix):
        """Create valuation sheet"""
        ws = self.wb.create_sheet(f'{prefix}_Valuation')
        ws['A1'] = 'VALUATION'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add valuation assumptions
        row = 3
        try:
            ev = scenario.exit_valuation
            ws[f'A{row}'] = 'VALUATION ASSUMPTIONS'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            assumptions = [
                ('Exit Year', ev.exit_year, ''),
                ('Exit Multiple (EV/EBITDA)', ev.exit_multiple_ev_ebitda, 'x'),
                ('Terminal Growth Rate', ev.terminal_growth_rate_pct, '%'),
                ('Discount Rate (NPV)', ev.discount_rate_npv_pct, '%'),
                ('Target IRR', ev.target_irr_pct, '%'),
                ('Valuation Method', ev.valuation_method, ''),
            ]
            for label, value, unit in assumptions:
                self._add_assumption(ws, row, label, value, unit)
                row += 1
        except Exception:
            pass
        
        row += 2
        
        # Add calculated valuation metrics
        from .models import CalculatedStatement
        val_stmts = CalculatedStatement.objects.filter(
            scenario=scenario, statement_type='valuation'
        )
        if val_stmts.exists():
            ws[f'A{row}'] = 'CALCULATED METRICS'
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            for stmt in val_stmts:
                if stmt.values_by_period:
                    for key, val in stmt.values_by_period.items():
                        ws[f'A{row}'] = key
                        if isinstance(val, (int, float)):
                            ws[f'B{row}'] = float(val)
                            if '%' in key:
                                ws[f'B{row}'].number_format = '0.00"%"'
                            else:
                                ws[f'B{row}'].number_format = '#,##0.00'
                        else:
                            ws[f'B{row}'] = val
                        ws[f'B{row}'].font = Font(bold=True, size=12)
                        row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _populate_statement_sheet(self, ws, title, statements):
        """Helper to populate statement sheets"""
        row = 1
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        ws[f'A{row}'] = 'Line Item'
        first_stmt = statements.first() if hasattr(statements, 'first') else statements[0]
        periods = sorted(first_stmt.values_by_period.keys())
        col = 2
        for period in periods:
            ws.cell(row, col).value = period
            ws.cell(row, col).font = Font(bold=True)
            col += 1
        row += 1
        
        # Data
        for stmt in statements:
            ws.cell(row, 1).value = stmt.line_item
            col = 2
            for period in periods:
                value = stmt.values_by_period.get(period, 0)
                ws.cell(row, col).value = float(value)
                ws.cell(row, col).number_format = '#,##0.00'
                col += 1
            row += 1
    
    def _format_statement_sheet(self, ws, num_periods):
        """Apply professional formatting to statement sheets"""
        # Column widths
        ws.column_dimensions['A'].width = 35
        for i in range(2, num_periods + 2):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        # Freeze panes
        ws.freeze_panes = 'B4'
    
    def _add_assumption(self, ws, row, label, value, unit=''):
        """Add assumption row with proper formatting"""
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        
        # Blue text for inputs
        ws[f'B{row}'].font = Font(color=self.BLUE)