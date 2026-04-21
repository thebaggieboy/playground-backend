"""
Enhanced Django Views for Financial Modeling Application
Implements calculation engine, Excel export, template management
"""

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.http import FileResponse, HttpResponse
from django.db import transaction, models
from django.utils import timezone
from datetime import datetime
import io
import json


# ============================================================================
# PAGINATION
# ============================================================================

class StandardResultsPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

from .models import (
    FinancialModel, Scenario, CalculatedStatement, ModelTemplate, CalculationLog
)
from .serializers import (
    FinancialModelListSerializer, FinancialModelDetailSerializer,
    FinancialModelCreateSerializer, ScenarioListSerializer,
    ScenarioDetailSerializer, ScenarioCreateUpdateSerializer,
    CalculatedStatementSerializer, ModelTemplateSerializer,
    TemplateCreateFromScenarioSerializer, CalculationLogSerializer
)


# ============================================================================
# FINANCIAL MODEL VIEWSET
# ============================================================================

class FinancialModelViewSet(viewsets.ModelViewSet):
    """
    Handles CRUD for Financial Models
    """
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsPagination  # #12
    
    def get_queryset(self):
        return FinancialModel.objects.filter(owner=self.request.user).prefetch_related('scenarios')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return FinancialModelCreateSerializer
        elif self.action == 'list':
            return FinancialModelListSerializer
        return FinancialModelDetailSerializer
    
    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """
        Trigger model calculation for all scenarios
        POST /api/models/{id}/calculate/
        """
        model = self.get_object()
        
        # Prevent concurrent calculations
        if model.is_calculation_in_progress:
            return Response(
                {'error': 'Calculation already in progress'},
                status=status.HTTP_409_CONFLICT
            )
        
        try:
            # Mark calculation as in progress
            model.is_calculation_in_progress = True
            model.save()
            
            # Run calculation for all active scenarios — capture per-scenario failures (#11)
            from .calculation_engine import CalculationEngine
            engine = CalculationEngine()
            
            results = {}
            failed_steps = []

            for scenario in model.scenarios.filter(is_active=True):
                try:
                    result = engine.calculate_scenario(scenario, user=request.user)
                    results[scenario.name] = result
                except Exception as scenario_err:
                    failed_steps.append(scenario.name)
                    results[scenario.name] = {'error': str(scenario_err)}
            
            # Update model status
            model.is_calculation_in_progress = False
            model.last_calculated_at = timezone.now()
            model.calculation_error = ', '.join(failed_steps) if failed_steps else None
            model.save()
            
            return Response({
                'status': 'success' if not failed_steps else 'partial',
                'message': 'Model calculated successfully' if not failed_steps else f'{len(failed_steps)} scenario(s) had errors',
                'results': results,
                'failed_steps': failed_steps,  # #11 — surfaced to frontend
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            model.is_calculation_in_progress = False
            model.calculation_error = str(e)
            model.save()
            
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def save_as_template(self, request, pk=None):
        """
        Save current model as a reusable template
        POST /api/models/{id}/save_as_template/
        Body: { "name": "Template Name", "description": "...", "is_public": false }
        """
        model = self.get_object()
        
        # Get base case scenario
        try:
            base_scenario = model.scenarios.get(scenario_type='base')
        except Scenario.DoesNotExist:
            return Response(
                {'error': 'Base case scenario not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Use project name as template name if not provided
        template_name = request.data.get('name', f"{model.name} Template")
        description = request.data.get('description', f"Template created from {model.name}")
        is_public = request.data.get('is_public', False)
        
        serializer = TemplateCreateFromScenarioSerializer(
            data={
                'name': template_name,
                'description': description,
                'is_public': is_public,
                'scenario_id': base_scenario.id
            },
            context={'request': request}
        )
        
        if serializer.is_valid():
            template = serializer.save()
            return Response(
                ModelTemplateSerializer(template).data,
                status=status.HTTP_201_CREATED
            )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Export model to Excel with all scenarios
        GET /api/models/{id}/export_excel/
        """
        model = self.get_object()
        
        try:
            from .excel_export import ExcelExporter
            exporter = ExcelExporter()
            
            # Generate Excel file
            excel_buffer = exporter.export_model(model)
            
            # Create response
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{model.name}.xlsx"'
            
            return response
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """
        Export model to PDF (uses base case scenario)
        GET /api/models/{id}/export_pdf/
        """
        model = self.get_object()
        
        try:
            # Use the base case scenario, fallback to first active
            scenario = (
                model.scenarios.filter(scenario_type='base').first() or
                model.scenarios.filter(is_active=True).first() or
                model.scenarios.first()
            )
            
            if not scenario:
                return Response(
                    {'error': 'No scenarios found for this model'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from .pdf_export import PDFExporter
            exporter = PDFExporter()
            pdf_buffer = exporter.export_scenario(scenario)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="{model.name}_Report.pdf"'
            
            return response
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def parse_upload(self, request):
        """
        Parse an uploaded financial model file (Excel, CSV, PDF).
        POST /api/models/parse_upload/
        Body: multipart/form-data with 'file' field
        Returns structured sheet data for frontend viewer.
        """
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        filename = uploaded_file.name
        file_size = uploaded_file.size
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        
        MAX_PREVIEW_ROWS = 200  # Limit rows for performance
        
        try:
            sheets = []
            total_cells = 0
            detected_type = 'Unknown'
            
            if ext in ('xlsx', 'xlsm', 'xls'):
                import openpyxl
                wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                detected_type = 'Excel Workbook'
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_data = []
                    headers = []
                    max_col = 0
                    row_count = 0
                    
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        row_count += 1
                        # Convert row to serializable list
                        clean_row = []
                        for cell in row:
                            if cell is None:
                                clean_row.append(None)
                            elif isinstance(cell, (int, float)):
                                clean_row.append(round(float(cell), 4) if isinstance(cell, float) else cell)
                            else:
                                clean_row.append(str(cell)[:200])  # Truncate long strings
                        
                        max_col = max(max_col, len(clean_row))
                        total_cells += len([c for c in clean_row if c is not None])
                        
                        if row_idx == 0:
                            headers = [str(c) if c else f'Col {i+1}' for i, c in enumerate(clean_row)]
                        elif row_idx <= MAX_PREVIEW_ROWS:
                            rows_data.append(clean_row)
                    
                    sheets.append({
                        'name': sheet_name,
                        'headers': headers[:50],  # Limit columns
                        'rows': [r[:50] for r in rows_data],
                        'totalRows': row_count,
                        'totalCols': min(max_col, 50),
                    })
                
                wb.close()
            
            elif ext == 'csv':
                import csv
                detected_type = 'CSV File'
                content = uploaded_file.read().decode('utf-8-sig', errors='replace')
                reader = csv.reader(content.splitlines())
                
                headers = []
                rows_data = []
                row_count = 0
                max_col = 0
                
                for row_idx, row in enumerate(reader):
                    row_count += 1
                    clean_row = []
                    for cell in row:
                        try:
                            val = float(cell.replace(',', ''))
                            clean_row.append(round(val, 4))
                        except (ValueError, AttributeError):
                            clean_row.append(cell[:200] if cell else None)
                    
                    max_col = max(max_col, len(clean_row))
                    total_cells += len([c for c in clean_row if c is not None])
                    
                    if row_idx == 0:
                        headers = [str(c) if c else f'Col {i+1}' for i, c in enumerate(clean_row)]
                    elif row_idx <= MAX_PREVIEW_ROWS:
                        rows_data.append(clean_row)
                
                sheets.append({
                    'name': 'Sheet1',
                    'headers': headers[:50],
                    'rows': [r[:50] for r in rows_data],
                    'totalRows': row_count,
                    'totalCols': min(max_col, 50),
                })
            
            elif ext == 'pdf':
                detected_type = 'PDF Document'
                # Basic PDF text extraction
                sheets.append({
                    'name': 'Document',
                    'headers': ['Content'],
                    'rows': [['PDF preview is not available. Please export as Excel for full data viewing.']],
                    'totalRows': 1,
                    'totalCols': 1,
                })
                total_cells = 1
            
            else:
                return Response({
                    'error': f'Unsupported file type: .{ext}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            return Response({
                'filename': filename,
                'fileSize': file_size,
                'sheets': sheets,
                'summary': {
                    'totalSheets': len(sheets),
                    'totalCells': total_cells,
                    'detectedType': detected_type,
                }
            })
        
        except Exception as e:
            return Response({
                'error': f'Failed to parse file: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# SCENARIO VIEWSET
# ============================================================================

class ScenarioViewSet(viewsets.ModelViewSet):
    """
    Handles CRUD for Scenarios with complete input data
    """
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsPagination  # #12
    
    def get_queryset(self):
        return Scenario.objects.filter(model__owner=self.request.user)
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ScenarioCreateUpdateSerializer
        elif self.action == 'list':
            # Allow analytics page to request full detail via ?detail=true
            if self.request.query_params.get('detail') == 'true':
                return ScenarioDetailSerializer
            return ScenarioListSerializer
        return ScenarioDetailSerializer
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """
        Duplicate scenario with all inputs
        POST /api/scenarios/{id}/duplicate/
        Body: { "name": "New Scenario Name" }
        """
        original = self.get_object()
        new_name = request.data.get('name', f"{original.name} - Copy")
        
        try:
            with transaction.atomic():
                # Serialize original scenario data
                original_data = ScenarioDetailSerializer(original).data
                
                # Remove id and set new name
                original_data.pop('id', None)
                original_data.pop('created_at', None)
                original_data['name'] = new_name
                original_data['scenario_type'] = 'custom'
                original_data['model'] = original.model.id
                
                # Create new scenario
                serializer = ScenarioCreateUpdateSerializer(
                    data=original_data,
                    context={'request': request}
                )
                
                if serializer.is_valid():
                    new_scenario = serializer.save()
                    return Response(
                        ScenarioDetailSerializer(new_scenario).data,
                        status=status.HTTP_201_CREATED
                    )
                
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def calculate(self, request, pk=None):
        """
        Calculate this specific scenario
        GET /api/scenarios/{id}/calculate/
        """
        scenario = self.get_object()
        
        try:
            from .calculation_engine import CalculationEngine
            engine = CalculationEngine()
            result = engine.calculate_scenario(scenario, user=request.user)
            
            return Response({
                'status': 'success',
                'result': result
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def sensitivity(self, request, pk=None):
        """
        Run a fast, memory-only calculation with overrides for Sensitivity Analysis.
        POST /api/scenarios/{id}/sensitivity/
        Body: {
            "revenue_growth_adj": 0.05,
            "opex_margin_adj": -0.02,
            ...
        }
        """
        scenario = self.get_object()
        overrides = request.data
        
        try:
            from .calculation_engine import CalculationEngine
            engine = CalculationEngine()
            result = engine.calculate_scenario(
                scenario, 
                user=request.user, 
                save_results=False, 
                overrides=overrides
            )
            
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Export scenario to Excel
        GET /api/scenarios/{id}/export_excel/
        """
        scenario = self.get_object()
        
        try:
            from .excel_export import ExcelExporter
            exporter = ExcelExporter()
            
            # Generate Excel file for this scenario
            excel_buffer = exporter.export_scenario(scenario)
            
            # Create response
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{scenario.model.name}_{scenario.name}.xlsx"'
            
            return response
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """
        Export scenario to PDF
        GET /api/scenarios/{id}/export_pdf/
        """
        scenario = self.get_object()
        
        try:
            from .pdf_export import PDFExporter
            exporter = PDFExporter()
            
            # Generate PDF
            pdf_buffer = exporter.export_scenario(scenario)
            
            # Create response
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="{scenario.model.name}_{scenario.name}.pdf"'
            
            return response
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# RESULTS VIEWSET
# ============================================================================

class CalculatedStatementViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only access to calculated statements
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = CalculatedStatementSerializer
    
    def get_queryset(self):
        return CalculatedStatement.objects.filter(
            scenario__model__owner=self.request.user
        )
    
    @action(detail=False, methods=['get'])
    def by_scenario(self, request):
        """
        Get all statements for a specific scenario
        GET /api/results/by_scenario/?scenario_id=1
        """
        scenario_id = request.query_params.get('scenario_id')
        
        if not scenario_id:
            return Response(
                {'error': 'scenario_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        statements = self.get_queryset().filter(scenario_id=scenario_id)
        
        # Group by statement type
        grouped = {}
        for stmt in statements:
            stmt_type = stmt.get_statement_type_display()
            if stmt_type not in grouped:
                grouped[stmt_type] = []
            grouped[stmt_type].append(self.get_serializer(stmt).data)
        
        return Response(grouped)


# ============================================================================
# TEMPLATE VIEWSET
# ============================================================================

class ModelTemplateViewSet(viewsets.ModelViewSet):
    """
    Manage reusable templates
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ModelTemplateSerializer
    
    def get_queryset(self):
        # Users see their own templates + public templates + system templates
        return ModelTemplate.objects.filter(
            models.Q(created_by=self.request.user) |
            models.Q(is_public=True) |
            models.Q(is_system_template=True)
        ).distinct()
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def apply_to_scenario(self, request, pk=None):
        """
        Apply template to an existing scenario
        POST /api/templates/{id}/apply_to_scenario/
        Body: { "scenario_id": 1 }
        """
        template = self.get_object()
        scenario_id = request.data.get('scenario_id')
        
        if not scenario_id:
            return Response(
                {'error': 'scenario_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            scenario = Scenario.objects.get(
                id=scenario_id,
                model__owner=request.user
            )
        except Scenario.DoesNotExist:
            return Response(
                {'error': 'Scenario not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            with transaction.atomic():
                # Load template data and apply to scenario
                template_data = template.template_data.copy()
                template_data['model'] = scenario.model.id
                template_data.pop('id', None)
                template_data.pop('created_at', None)
                
                serializer = ScenarioCreateUpdateSerializer(
                    scenario,
                    data=template_data,
                    partial=True,
                    context={'request': request}
                )
                
                if serializer.is_valid():
                    serializer.save()
                    return Response(
                        ScenarioDetailSerializer(scenario).data,
                        status=status.HTTP_200_OK
                    )
                
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# CALCULATION LOG VIEWSET
# ============================================================================

class CalculationLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    View calculation history and logs
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = CalculationLogSerializer
    
    def get_queryset(self):
        return CalculationLog.objects.filter(
            scenario__model__owner=self.request.user
        ).select_related('scenario', 'triggered_by')