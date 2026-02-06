import openpyxl
import os
from decimal import Decimal
from django.conf import settings
from .models import ModelTemplate, TemplateMapping, Assumption, HistoricalData, Scenario, FinancialModel

# Define the sheet and columns where the template mapping metadata is stored in the Excel file
MAPPING_SHEET_NAME = 'Template_Mapping_Meta'
VARIABLE_COL = 'A' # Column containing the variable_name (e.g., 'RevenueGrowthRate')
REFERENCE_COL = 'B' # Column containing the Excel cell reference (e.g., 'Assumptions!B10')
TYPE_COL = 'C' # Column containing the data type (e.g., 'Input', 'Historical_IS')
GUIDELINE_COL = 'D' # Column containing the guideline text

def parse_excel_template(template: ModelTemplate):
    """
    Reads the metadata sheet from the template file and creates TemplateMapping records.
    Assumes the template file is stored in MEDIA_ROOT/model_templates/.
    """
    # Build the full path to the file
    file_path = os.path.join(settings.MEDIA_ROOT, template.data_source_file.name)
    
    # Clear any existing mappings for this template before creating new ones
    TemplateMapping.objects.filter(template=template).delete()

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if MAPPING_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Required mapping sheet '{MAPPING_SHEET_NAME}' not found in the Excel file.")

        sheet = workbook[MAPPING_SHEET_NAME]
        mappings = []
        
        # Iterate through rows, starting from the second row (assuming row 1 is headers)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if not row[0].value: continue # Skip if the variable name is empty
            
            # Extract values based on column definitions
            variable = row[openpyxl.utils.column_index_from_string(VARIABLE_COL)-1].value
            reference = row[openpyxl.utils.column_index_from_string(REFERENCE_COL)-1].value
            data_type = row[openpyxl.utils.column_index_from_string(TYPE_COL)-1].value
            guideline = row[openpyxl.utils.column_index_from_string(GUIDELINE_COL)-1].value or ""

            mappings.append(
                TemplateMapping(
                    template=template,
                    variable_name=variable,
                    excel_reference=reference,
                    data_type=data_type,
                    guideline=guideline
                )
            )
            
        TemplateMapping.objects.bulk_create(mappings)

    except Exception as e:
        raise Exception(f"Failed to parse Excel file: {e}")


def hydrate_model_data(new_model: FinancialModel, base_scenario: Scenario, template: ModelTemplate):
    """
    Pulls data from the Excel file referenced by TemplateMapping and populates
    HistoricalData and initial Assumption records for a newly created model.
    """
    file_path = os.path.join(settings.MEDIA_ROOT, template.data_source_file.name)
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    mappings = TemplateMapping.objects.filter(template=template)
    
    historical_records = []
    assumption_records = []

    for mapping in mappings:
        try:
            # Parse the reference (e.g., 'SheetName!B10')
            sheet_name, cell_ref = mapping.excel_reference.split('!')
            sheet = workbook[sheet_name]
            cell_value = sheet[cell_ref].value
            
            # Convert value to Decimal for financial accuracy
            value = Decimal(str(cell_value)) if cell_value is not None else Decimal(0)
            
            # Use data_type to determine where to save the data
            if mapping.data_type.startswith('Historical'):
                # Historical data usually spans multiple periods, but for simplicity, we treat it as a single initial value.
                # In a real model, this would require advanced range parsing.
                historical_records.append(
                    HistoricalData(
                        model=new_model,
                        line_item=mapping.variable_name,
                        period='2024-Q4', # Use a fixed initial period
                        value=value
                    )
                )
            elif mapping.data_type == 'Input':
                # Input data is created as the starting assumption for the base scenario
                assumption_records.append(
                    Assumption(
                        scenario=base_scenario,
                        variable_name=mapping.variable_name,
                        period='2025-Q1', # Use a fixed starting forecast period
                        value=value
                    )
                )
        except Exception as e:
            print(f"Skipping mapping {mapping.variable_name}: Error - {e}")

    HistoricalData.objects.bulk_create(historical_records)
    Assumption.objects.bulk_create(assumption_records)